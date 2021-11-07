from PySimpleGUI.PySimpleGUI import Column
import openpyxl
import pandas as pd
import numpy as np
import datetime

from user import User

def addUser(User):
    book = openpyxl.load_workbook('user.xlsx')
    sheet = book.worksheets[0]
    endrow = sheet.max_row + 1
    
    sheet.cell(row=endrow, column=1).value = User.email
    sheet.cell(row=endrow, column=2).value = User.pw
    sheet.cell(row=endrow, column=3).value = User.name
    sheet.cell(row=endrow, column=4).value = User.gender
    sheet.cell(row=endrow, column=5).value = User.birth_year
    sheet.cell(row=endrow, column=6).value = User.birth_month
    sheet.cell(row=endrow, column=7).value = User.birth_date

    book.save('user.xlsx')
    book.close()

def getUser(id):
    book = openpyxl.load_workbook('user.xlsx')
    sheet = book.worksheets[0]

    for i in range(1, sheet.max_row+1):
        rownum = i
        if sheet.cell(row=rownum, column=1).value == id:
            userinfo = {
                'email': sheet.cell(row=rownum, column=1).value,
                'pw': sheet.cell(row=rownum, column=2).value,
                'name': sheet.cell(row=rownum, column=3).value,
                'gender': sheet.cell(row=rownum, column=4).value,
                'birth_year': sheet.cell(row=rownum, column=5).value,
                'birth_month': sheet.cell(row=rownum, column=6).value,
                'birth_date': sheet.cell(row=rownum, column=7).value,
            }
            break
    return userinfo

def getRow():
    book = openpyxl.load_workbook('user.xlsx')
    sheet = book.worksheets[1]
    row_num = sheet.max_row
    return row_num

def getOrder(num):
    book = openpyxl.load_workbook('user.xlsx')
    sheet = book.worksheets[1]
    ret = {}

    ret['room'] = sheet.cell(row=num, column=1).value
    ret['checkin_hour'] = sheet.cell(row=num, column=2).value
    ret['checkin_min'] = sheet.cell(row=num, column=3).value
    ret['checkout_hour'] = sheet.cell(row=num, column=4).value
    ret['checkout_min'] = sheet.cell(row=num, column=5).value

    return ret

def getSendTo():
    book = openpyxl.load_workbook('user.xlsx')
    sheet = book.worksheets[0]
    value = sheet.cell(row=sheet.max_row, column=1).value
    return value

def authenticate(email, pw):
    book = openpyxl.load_workbook('user.xlsx')
    sheet = book.worksheets[0]

    ret = None

    for i in range(1, sheet.max_row+1):
        rownum = i
        if sheet.cell(row=rownum, column=1).value == email:
            if sheet.cell(row=rownum, column=2).value == pw:
                ret = True
                break
            else:
                ret = False
        else:
            ret = False

    book.close()
    return ret

def setReserve(room, in_hour, in_min, out_hour, out_min):
    book = openpyxl.load_workbook('user.xlsx')
    sheet = book.worksheets[1]
    rownum = sheet.max_row + 1

    sheet.cell(row=rownum, column=1).value = room
    sheet.cell(row=rownum, column=2).value = in_hour
    sheet.cell(row=rownum, column=3).value = in_min
    sheet.cell(row=rownum, column=4).value = out_hour
    sheet.cell(row=rownum, column=5).value = out_min

    book.save('user.xlsx')
    book.close()